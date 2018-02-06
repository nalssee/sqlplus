#!/usr/bin/env python

import os
import csv
import argparse
from openpyxl import load_workbook


parser = argparse.ArgumentParser()
parser.add_argument("filename", type=str,
                    help="Excel file to multiple CSVs")


args = parser.parse_args()
filename = args.filename
fname = os.path.splitext(filename)[0]

workbook = load_workbook(filename)
for sheet in workbook.sheetnames:
    with open(fname + '_' + sheet + '.csv', 'w',
              newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        for row in workbook[sheet].iter_rows():
            w.writerow(c.value for c in row)


